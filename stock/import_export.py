# stock/import_export.py
"""Export et import du catalogue produits au format Excel (.xlsx).

Le fichier exporté et le fichier attendu à l'import ont EXACTEMENT les mêmes
colonnes : on peut exporter, modifier dans Excel, puis réimporter tel quel
pour mettre à jour le catalogue en masse."""
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

COLONNES = [
    'Référence', 'Nom', 'Marque', 'Catégorie', 'Fournisseur',
    'Prix achat', 'Prix vente', 'Marge (%)', 'Stock actuel',
    'Seuil minimum', 'Seuil maximum', 'Code-barres', 'Description',
]


def exporter_produits_xlsx(produits):
    """produits : queryset ou liste de Produit (avec select_related déjà
    appliqué idéalement). Retourne un buffer BytesIO prêt à être servi."""
    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "Produits"

    feuille.append(COLONNES)
    for cellule in feuille[1]:
        cellule.font = Font(bold=True, color="FFFFFF")
        cellule.fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")

    for produit in produits:
        feuille.append([
            produit.reference,
            produit.nom,
            produit.marque.nom if produit.marque_id else '',
            produit.categorie.nom if produit.categorie_id else '',
            produit.fournisseur.nom if produit.fournisseur_id else '',
            float(produit.prix_achat),
            float(produit.prix_vente),
            produit.taux_marge if produit.taux_marge is not None else '',
            produit.quantite_stock,
            produit.seuil_alerte,
            produit.seuil_max if produit.seuil_max is not None else '',
            produit.code_barres,
            produit.description,
        ])

    largeurs = [14, 28, 16, 16, 18, 12, 12, 11, 12, 13, 13, 16, 30]
    for i, largeur in enumerate(largeurs, start=1):
        feuille.column_dimensions[get_column_letter(i)].width = largeur

    tampon = io.BytesIO()
    classeur.save(tampon)
    tampon.seek(0)
    return tampon


class ResultatImport:
    def __init__(self):
        self.crees = 0
        self.modifies = 0
        self.erreurs = []

    @property
    def total_traite(self):
        return self.crees + self.modifies

    @property
    def a_reussi(self):
        return self.total_traite > 0 or not self.erreurs


def importer_produits_xlsx(fichier, Produit, Marque, Categorie, Fournisseur):
    """fichier : objet fichier Django (request.FILES['fichier']).
    Les modèles sont passés en paramètre plutôt qu'importés directement pour
    que cette fonction reste testable indépendamment de l'app Django."""
    resultat = ResultatImport()

    try:
        classeur = load_workbook(fichier, data_only=True)
    except Exception as erreur:
        resultat.erreurs.append(f"Fichier illisible : le fichier doit être un .xlsx valide ({erreur}).")
        return resultat

    feuille = classeur.active
    lignes = list(feuille.iter_rows(min_row=2, values_only=True))

    for numero_ligne, ligne in enumerate(lignes, start=2):
        if ligne is None or all(v in (None, '') for v in ligne):
            continue  # ligne vide, on l'ignore silencieusement

        try:
            (reference, nom, nom_marque, nom_categorie, nom_fournisseur,
             prix_achat, prix_vente, _marge_ignoree, _stock_ignore,
             seuil_min, seuil_max, code_barres, description) = (list(ligne) + [None] * 13)[:13]

            if not nom or str(nom).strip() == '':
                resultat.erreurs.append(f"Ligne {numero_ligne} : nom du produit manquant, ligne ignorée.")
                continue
            if prix_achat in (None, '') or prix_vente in (None, ''):
                resultat.erreurs.append(f"Ligne {numero_ligne} ({nom}) : prix d'achat/vente manquant, ligne ignorée.")
                continue

            marque = None
            if nom_marque and str(nom_marque).strip():
                marque, _ = Marque.objects.get_or_create(nom=str(nom_marque).strip())

            categorie = None
            if nom_categorie and str(nom_categorie).strip():
                categorie, _ = Categorie.objects.get_or_create(nom=str(nom_categorie).strip())

            fournisseur = None
            if nom_fournisseur and str(nom_fournisseur).strip():
                # Volontairement PAS de get_or_create ici : un fournisseur a besoin
                # d'une fiche à part (contact, adresse...) qu'on ne veut pas créer
                # à la volée depuis une ligne de produit avec un simple nom.
                fournisseur = Fournisseur.objects.filter(nom__iexact=str(nom_fournisseur).strip()).first()
                if fournisseur is None:
                    resultat.erreurs.append(
                        f"Ligne {numero_ligne} ({nom}) : fournisseur « {nom_fournisseur} » introuvable, "
                        f"produit importé sans fournisseur associé."
                    )

            reference = str(reference).strip() if reference else ''
            produit = Produit.objects.filter(reference=reference).first() if reference else None
            est_creation = produit is None
            if produit is None:
                produit = Produit(reference=reference)

            produit.nom = str(nom).strip()
            produit.marque = marque
            produit.categorie = categorie
            produit.fournisseur = fournisseur
            produit.prix_achat = prix_achat
            produit.prix_vente = prix_vente
            produit.seuil_alerte = seuil_min if seuil_min not in (None, '') else 5
            produit.seuil_max = seuil_max if seuil_max not in (None, '') else None
            produit.code_barres = str(code_barres).strip() if code_barres else ''
            produit.description = str(description).strip() if description else ''
            # La quantité en stock n'est JAMAIS écrasée par un import : elle ne
            # doit bouger que via de vrais MouvementStock (achat, vente,
            # inventaire), jamais par une valeur tapée dans un tableur.
            produit.save()

            if est_creation:
                resultat.crees += 1
            else:
                resultat.modifies += 1

        except Exception as erreur:
            resultat.erreurs.append(f"Ligne {numero_ligne} : erreur inattendue ({erreur}), ligne ignorée.")

    return resultat
